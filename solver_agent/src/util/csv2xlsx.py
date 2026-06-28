import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Column whose equal consecutive values get merged into one grouped cell.
GROUP_COLUMN_NAME = "peripheral"



def read_csv(path):
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        raise ValueError("CSV is empty")

    header = rows[0]
    data = rows[1:]
    return header, data


def autosize_columns(ws, max_width=45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0

        for cell in col:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def apply_group_border(ws, start_row, end_row, start_col, end_col):
    thick = Side(style="medium", color="000000")
    thin = Side(style="thin", color="D9D9D9")

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)

            left = thick if col == start_col else thin
            right = thick if col == end_col else thin
            top = thick if row == start_row else thin
            bottom = thick if row == end_row else thin

            cell.border = Border(
                left=left,
                right=right,
                top=top,
                bottom=bottom
            )


def merge_same_peripheral(ws, header, data_start_row=2):
    if GROUP_COLUMN_NAME not in header:
        raise ValueError(f"Cannot find column: {GROUP_COLUMN_NAME}")

    group_col = header.index(GROUP_COLUMN_NAME) + 1
    max_col = len(header)
    max_row = ws.max_row

    current_value = None
    group_start = data_start_row

    for row in range(data_start_row, max_row + 2):
        value = ws.cell(row=row, column=group_col).value if row <= max_row else None

        if row == data_start_row:
            current_value = value
            group_start = row
            continue

        if value != current_value:
            group_end = row - 1

            if current_value is not None and group_end >= group_start:
                # 同一個 peripheral 超過一列才 merge
                if group_end > group_start:
                    ws.merge_cells(
                        start_row=group_start,
                        start_column=group_col,
                        end_row=group_end,
                        end_column=group_col
                    )

                merged_cell = ws.cell(row=group_start, column=group_col)
                merged_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                merged_cell.font = Font(bold=True)

                # 對整個 peripheral group 加大框
                apply_group_border(
                    ws,
                    start_row=group_start,
                    end_row=group_end,
                    start_col=1,
                    end_col=max_col
                )

            current_value = value
            group_start = row


def csv_to_grouped_xlsx(input_csv, output_xlsx):
    header, data = read_csv(input_csv)

    wb = Workbook()
    ws = wb.active
    ws.title = "plan"

    # 寫入 header
    ws.append(header)

    # 寫入資料
    for row in data:
        ws.append(row)

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body style
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 合併同 peripheral 的欄位，並加 group 外框
    merge_same_peripheral(ws, header)

    # 欄寬
    autosize_columns(ws)

    # 凍結 header
    ws.freeze_panes = "A2"

    # 開啟 filter
    ws.auto_filter.ref = ws.dimensions

    # row height
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24

    wb.save(output_xlsx)
    print(f"Saved: {output_xlsx}")


